"""Create an Excel dashboard from collected YES24 CSV data."""

from __future__ import annotations

import html
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_PATH = BASE_DIR / "data" / "yes24_it_mobile_bestsellers.csv"
OUTPUT_DIR = BASE_DIR / "outputs" / "yes24_dashboard"
OUTPUT_PATH = OUTPUT_DIR / "yes24_books_dashboard.xlsx"


COLORS = {
    "navy": "1F4E78",
    "blue": "5B9BD5",
    "light_blue": "D9EAF7",
    "teal": "2F75B5",
    "green": "70AD47",
    "orange": "F4B183",
    "gray": "F2F2F2",
    "dark_gray": "595959",
    "white": "FFFFFF",
    "border": "D9E2F3",
}


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    text = html.unescape(str(value))
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def to_number(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series.astype(str).str.replace(",", "", regex=False).str.strip(), errors="coerce")


def load_yes24_data() -> pd.DataFrame:
    df = pd.read_csv(DATA_PATH)
    numeric_cols = ["순위", "판매가(원)", "정가(원)", "할인율(%)", "평점", "리뷰수", "판매지수"]
    for col in numeric_cols:
        df[col] = to_number(df[col])
    df["내용"] = df["내용"].map(clean_text)
    df["출간연도"] = df["출간일"].astype(str).str.extract(r"(\d{4})").astype(float)
    return df


def set_title(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=20, color=COLORS["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=10, color=COLORS["dark_gray"])
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["gray"])
    ws.row_dimensions[2].height = 22


def style_range(ws, cell_range: str, fill: str | None = None, bold: bool = False, font_color: str = "000000") -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.font = Font(name="Arial", bold=bold, color=font_color)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = Border(bottom=Side(style="thin", color=COLORS["border"]))
            cell.alignment = Alignment(vertical="center")


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start_col):
        ws.cell(start_row, col_idx, header)
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, start_col):
            ws.cell(row_idx, col_idx, None if pd.isna(value) else value)


def add_table(ws, name: str, start_row: int, start_col: int, rows: int, cols: int) -> None:
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + cols - 1)}{start_row + rows - 1}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)


def prepare_analysis(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    publisher_counts = df["출판사"].value_counts().head(10).reset_index()
    publisher_counts.columns = ["출판사", "도서수"]

    publisher_sales = (
        df.groupby("출판사", as_index=False)
        .agg(평균판매지수=("판매지수", "mean"), 도서수=("도서명", "count"))
        .query("도서수 >= 2")
        .sort_values("평균판매지수", ascending=False)
        .head(10)
    )

    top_sales = df.sort_values("판매지수", ascending=False).head(10)[["도서명", "판매지수", "평점", "리뷰수", "출판사"]]
    top_sales["도서명"] = top_sales["도서명"].str.slice(0, 28)

    rating_bins = pd.cut(df["평점"], bins=[0, 7, 8, 9, 9.5, 10], include_lowest=True)
    rating_dist = rating_bins.value_counts().sort_index().reset_index()
    rating_dist.columns = ["평점구간", "도서수"]
    rating_dist["평점구간"] = rating_dist["평점구간"].astype(str)

    price_bins = pd.cut(
        df["판매가(원)"],
        bins=[0, 10000, 20000, 30000, 40000, 1000000],
        labels=["1만원 이하", "1~2만원", "2~3만원", "3~4만원", "4만원 초과"],
        include_lowest=True,
    )
    price_dist = price_bins.value_counts().sort_index().reset_index()
    price_dist.columns = ["가격구간", "도서수"]
    price_dist["가격구간"] = price_dist["가격구간"].astype(str)

    yearly = (
        df.dropna(subset=["출간연도"])
        .groupby("출간연도", as_index=False)
        .agg(도서수=("도서명", "count"), 평균평점=("평점", "mean"), 평균판매지수=("판매지수", "mean"))
        .sort_values("출간연도")
    )
    yearly["출간연도"] = yearly["출간연도"].astype(int).astype(str)

    return {
        "publisher_counts": publisher_counts,
        "publisher_sales": publisher_sales,
        "top_sales": top_sales,
        "rating_dist": rating_dist,
        "price_dist": price_dist,
        "yearly": yearly,
    }


def setup_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Data")
    write_dataframe(ws, df)
    add_table(ws, "BooksData", 1, 1, len(df) + 1, len(df.columns))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 8,
        "B": 13,
        "C": 42,
        "D": 18,
        "E": 18,
        "F": 14,
        "G": 12,
        "H": 12,
        "I": 11,
        "J": 9,
        "K": 9,
        "L": 12,
        "M": 42,
        "N": 80,
        "O": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color=COLORS["white"])
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.alignment = Alignment(horizontal="center")

    for col in ["G", "H", "L"]:
        for cell in ws[col][1:]:
            cell.number_format = "#,##0"
    for col in ["I", "J"]:
        for cell in ws[col][1:]:
            cell.number_format = "0.0"
    for cell in ws["N"][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def setup_analysis_sheet(wb: Workbook, analysis: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    set_title(ws, "YES24 도서 분석 테이블", "대시보드 차트의 원천 데이터입니다.")

    sections = [
        ("A4", "출판사별 도서 수 Top 10", analysis["publisher_counts"]),
        ("D4", "출판사별 평균 판매지수 Top 10", analysis["publisher_sales"]),
        ("H4", "판매지수 Top 10 도서", analysis["top_sales"]),
        ("A20", "평점 분포", analysis["rating_dist"]),
        ("D20", "판매가 분포", analysis["price_dist"]),
        ("G20", "출간연도별 추이", analysis["yearly"]),
    ]

    for anchor, title, frame in sections:
        col = ws[anchor].column
        row = ws[anchor].row
        ws.cell(row, col, title)
        ws.cell(row, col).font = Font(name="Arial", bold=True, size=12, color=COLORS["navy"])
        write_dataframe(ws, frame, row + 1, col)
        last_col = col + len(frame.columns) - 1
        header_range = f"{get_column_letter(col)}{row + 1}:{get_column_letter(last_col)}{row + 1}"
        style_range(ws, header_range, fill=COLORS["light_blue"], bold=True, font_color=COLORS["navy"])

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["H"].width = 34
    ws.column_dimensions["L"].width = 18

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", bold=cell.font.bold, size=cell.font.sz or 10, color=cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else "000000")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.0" if "평균" in str(ws.cell(cell.row - (cell.row > 1), cell.column).value) else "#,##0"


def kpi_card(ws, cell: str, title: str, formula: str, number_format: str) -> None:
    start = ws[cell]
    row, col = start.row, start.column
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.cell(row, col, title)
    ws.cell(row, col).font = Font(name="Arial", bold=True, size=10, color=COLORS["dark_gray"])
    ws.cell(row, col).fill = PatternFill("solid", fgColor=COLORS["gray"])
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
    ws.cell(row + 1, col, formula)
    ws.cell(row + 1, col).font = Font(name="Arial", bold=True, size=18, color=COLORS["navy"])
    ws.cell(row + 1, col).fill = PatternFill("solid", fgColor=COLORS["white"])
    ws.cell(row + 1, col).number_format = number_format
    ws.cell(row + 1, col).alignment = Alignment(horizontal="center", vertical="center")
    for r in range(row, row + 3):
        for c in range(col, col + 2):
            ws.cell(r, c).border = Border(
                top=Side(style="thin", color=COLORS["border"]),
                bottom=Side(style="thin", color=COLORS["border"]),
                left=Side(style="thin", color=COLORS["border"]),
                right=Side(style="thin", color=COLORS["border"]),
            )


def add_dashboard_charts(ws) -> None:
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.style = 10
    chart1.title = "출판사별 도서 수 Top 10"
    chart1.y_axis.title = "출판사"
    chart1.x_axis.title = "도서 수"
    chart1.add_data(Reference(ws, min_col=12, min_row=6, max_row=15), titles_from_data=False)
    chart1.set_categories(Reference(ws, min_col=11, min_row=6, max_row=15))
    chart1.height = 8
    chart1.width = 13
    chart1.legend = None
    ws.add_chart(chart1, "A9")

    chart2 = BarChart()
    chart2.style = 11
    chart2.title = "판매지수 Top 10 도서"
    chart2.y_axis.title = "판매지수"
    chart2.add_data(Reference(ws, min_col=15, min_row=6, max_row=15), titles_from_data=False)
    chart2.set_categories(Reference(ws, min_col=14, min_row=6, max_row=15))
    chart2.height = 8
    chart2.width = 16
    chart2.legend = None
    ws.add_chart(chart2, "H9")

    chart3 = PieChart()
    chart3.title = "판매가 구간별 도서 비중"
    chart3.add_data(Reference(ws, min_col=18, min_row=6, max_row=10), titles_from_data=False)
    chart3.set_categories(Reference(ws, min_col=17, min_row=6, max_row=10))
    chart3.dataLabels = DataLabelList()
    chart3.dataLabels.showPercent = True
    chart3.height = 7
    chart3.width = 9
    ws.add_chart(chart3, "A26")

    chart4 = LineChart()
    chart4.title = "출간연도별 도서 수"
    chart4.y_axis.title = "도서 수"
    chart4.x_axis.title = "출간연도"
    chart4.add_data(Reference(ws, min_col=21, min_row=5, max_row=18), titles_from_data=True)
    chart4.set_categories(Reference(ws, min_col=20, min_row=6, max_row=18))
    chart4.height = 7
    chart4.width = 13
    ws.add_chart(chart4, "H26")


def setup_dashboard_sheet(wb: Workbook, df: pd.DataFrame, analysis: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    set_title(ws, "YES24 IT/모바일 베스트셀러 대시보드", f"Source: {DATA_PATH.name} | 총 {len(df):,}권")

    for col in range(1, 22):
        ws.column_dimensions[get_column_letter(col)].width = 13
    for row in range(1, 40):
        ws.row_dimensions[row].height = 20

    last_data_row = len(df) + 1
    kpi_card(ws, "A4", "도서 수", '=COUNTA(Data!$C$2:$C$1001)', '#,##0"권"')
    kpi_card(ws, "C4", "출판사 수", '=COUNTA(UNIQUE(Data!$E$2:$E$1001))', '#,##0"곳"')
    kpi_card(ws, "E4", "평균 평점", f'=AVERAGE(Data!$J$2:$J${last_data_row})', '0.00')
    kpi_card(ws, "G4", "평균 판매가", f'=AVERAGE(Data!$G$2:$G${last_data_row})', '#,##0"원"')
    kpi_card(ws, "I4", "평균 판매지수", f'=AVERAGE(Data!$L$2:$L${last_data_row})', '#,##0')

    helper_tables = [
        ("K5", analysis["publisher_counts"]),
        ("N5", analysis["top_sales"][["도서명", "판매지수"]]),
        ("Q5", analysis["price_dist"]),
        ("T5", analysis["yearly"][["출간연도", "도서수"]].tail(13)),
    ]
    for anchor, frame in helper_tables:
        write_dataframe(ws, frame, ws[anchor].row, ws[anchor].column)
        end_col = ws[anchor].column + len(frame.columns) - 1
        style_range(ws, f"{anchor}:{get_column_letter(end_col)}{ws[anchor].row}", fill=COLORS["light_blue"], bold=True, font_color=COLORS["navy"])

    ws.column_dimensions["M"].width = 34
    ws.column_dimensions["P"].width = 16
    ws.column_dimensions["S"].width = 12
    ws.column_dimensions["T"].width = 12

    add_dashboard_charts(ws)

    ws["A37"] = "대시보드 사용 팁: Data 시트에서 필터를 사용해 도서별 세부 정보를 확인하고, Analysis 시트에서 차트 원천 테이블을 검토할 수 있습니다."
    ws.merge_cells("A37:T37")
    ws["A37"].font = Font(name="Arial", italic=True, color=COLORS["dark_gray"])
    ws["A37"].fill = PatternFill("solid", fgColor=COLORS["gray"])

    ws.freeze_panes = "A4"
    ws.conditional_formatting.add(
        "O6:O15",
        ColorScaleRule(start_type="min", start_color="FCE4D6", mid_type="percentile", mid_value=50, mid_color="FFF2CC", end_type="max", end_color="70AD47"),
    )


def verify_workbook(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    required = {"Dashboard", "Analysis", "Data"}
    missing = required.difference(wb.sheetnames)
    if missing:
        raise RuntimeError(f"Missing sheets: {sorted(missing)}")

    formula_errors = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(err in cell.value for err in formula_errors):
                    raise RuntimeError(f"Formula error marker found at {ws.title}!{cell.coordinate}: {cell.value}")

    if len(wb["Dashboard"]._charts) < 4:
        raise RuntimeError("Dashboard has fewer than 4 charts")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df = load_yes24_data()
    analysis = prepare_analysis(df)

    wb = Workbook()
    wb.remove(wb.active)
    setup_dashboard_sheet(wb, df, analysis)
    setup_analysis_sheet(wb, analysis)
    setup_data_sheet(wb, df)

    wb.properties.title = "YES24 Books Dashboard"
    wb.properties.subject = "Exploratory data analysis for collected YES24 data"
    wb.properties.creator = "Codex"
    wb.save(OUTPUT_PATH)
    verify_workbook(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
